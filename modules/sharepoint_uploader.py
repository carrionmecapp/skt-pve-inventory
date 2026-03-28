"""
Módulo para subir archivos a SharePoint
"""

import requests
from datetime import datetime
from io import BytesIO

# Imports opcionales para SharePoint
try:
    import msal
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    SHAREPOINT_AVAILABLE = True
except ImportError:
    SHAREPOINT_AVAILABLE = False
    print("⚠️ Dependencias SharePoint no instaladas. Instala con: pip install msal openpyxl")


class SharePointUploader:
    """Clase para subir datos directamente a SharePoint"""
    
    def __init__(self, config_manager):
        if not SHAREPOINT_AVAILABLE:
            print("❌ SharePoint no disponible. Instala dependencias: pip install msal openpyxl")
            return
        
        self.config = config_manager.get_sharepoint_config()
        self.access_token = None
        
        # Variables de configuración
        self.tenant_id = self.config.get('tenant_id')
        self.client_id = self.config.get('client_id')
        self.client_secret = self.config.get('client_secret')
        self.sharepoint_url = self.config.get('sharepoint_url')
        self.site_name = self.config.get('site_name')
        self.library_name = self.config.get('library_name', 'Documents')
        self.default_folder = self.config.get('default_folder', 'Proxmox-VMs')
        self.default_filename = self.config.get('default_filename', 'inventario_proxmox_vms.xlsx')
    
    def is_configured(self):
        """Verificar si SharePoint está configurado"""
        return SHAREPOINT_AVAILABLE and bool(self.tenant_id and self.client_id and self.client_secret)
    
    def authenticate(self):
        """Autenticar con SharePoint usando MSAL"""
        if not SHAREPOINT_AVAILABLE:
            return False
        
        try:
            app = msal.ConfidentialClientApplication(
                client_id=self.client_id,
                client_credential=self.client_secret,
                authority=f"https://login.microsoftonline.com/{self.tenant_id}"
            )
            
            result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
            
            if "access_token" in result:
                self.access_token = result["access_token"]
                print("✓ Autenticación exitosa con SharePoint")
                return True
            else:
                print(f"❌ Error autenticación SharePoint: {result.get('error_description', 'Error desconocido')}")
                return False
                
        except Exception as e:
            print(f"❌ Error autenticando con SharePoint: {e}")
            return False
    
    def get_site_id(self):
        """Obtener ID del sitio SharePoint"""
        try:
            if self.site_name:
                # Sitio específico
                site_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_url.split('//')[-1]}:/sites/{self.site_name}"
            else:
                # Sitio raíz
                site_url = f"https://graph.microsoft.com/v1.0/sites/{self.sharepoint_url.split('//')[-1]}"
            
            headers = {'Authorization': f'Bearer {self.access_token}'}
            response = requests.get(site_url, headers=headers)
            
            if response.status_code == 200:
                return response.json()['id']
            else:
                print(f"❌ Error obteniendo sitio: {response.status_code}")
                return None
                
        except Exception as e:
            print(f"❌ Error obteniendo site ID: {e}")
            return None
    
    def create_folder_if_not_exists(self, site_id, folder_path):
        """Crear carpeta si no existe"""
        try:
            # Verificar si la carpeta existe
            folder_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{folder_path}"
            headers = {'Authorization': f'Bearer {self.access_token}'}
            response = requests.get(folder_url, headers=headers)
            
            if response.status_code == 404:
                # Crear carpeta
                create_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/children"
                data = {
                    "name": folder_path,
                    "folder": {}
                }
                response = requests.post(create_url, headers={**headers, 'Content-Type': 'application/json'}, 
                                       json=data)
                if response.status_code == 201:
                    print(f"✓ Carpeta '{folder_path}' creada")
                    
        except Exception as e:
            print(f"⚠️ Error creando carpeta: {e}")
    
    def download_existing_excel(self, site_id, full_path):
        """Descargar Excel existente desde SharePoint"""
        try:
            download_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{full_path}:/content"
            headers = {'Authorization': f'Bearer {self.access_token}'}
            
            response = requests.get(download_url, headers=headers)
            
            if response.status_code == 200:
                return BytesIO(response.content)
            else:
                print(f"⚠️ No se pudo descargar archivo existente: {response.status_code}")
                return None
                
        except Exception as e:
            print(f"⚠️ Error descargando archivo existente: {e}")
            return None
    
    def update_existing_excel_in_memory(self, inventory, site_id, full_path):
        """Actualizar Excel existente manteniendo estructura y tablas dinámicas"""
        if not SHAREPOINT_AVAILABLE:
            return BytesIO()
        
        # Intentar descargar archivo existente
        existing_file = self.download_existing_excel(site_id, full_path)
        
        if existing_file:
            try:
                print("📄 Actualizando archivo Excel existente...")
                # Cargar archivo existente
                wb = openpyxl.load_workbook(existing_file)
                
                # Buscar hoja de datos (priorizar "VM Inventory" o la primera hoja)
                if "VM Inventory" in wb.sheetnames:
                    ws = wb["VM Inventory"]
                    print("✓ Encontrada hoja 'VM Inventory'")
                elif "Datos" in wb.sheetnames:
                    ws = wb["Datos"]
                    print("✓ Encontrada hoja 'Datos'")
                else:
                    ws = wb.active
                    print(f"✓ Usando hoja activa: {ws.title}")
                
            except Exception as e:
                print(f"⚠️ Error cargando archivo existente, creando nuevo: {e}")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "VM Inventory"
                existing_file = None
        else:
            print("📄 Creando nuevo archivo Excel...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "VM Inventory"
            existing_file = None
        
        if not inventory:
            return BytesIO()
        
        # Headers ordenados
        headers = [
            'fecha_actualizacion', 'cluster_name', 'cluster_type', 'node', 'vmid', 'name', 'status',
            'cpu_cores', 'cpu_sockets', 'memory_gb', 'total_disk_gb', 'disk_count',
            'network_interfaces', 'ip_count', 'ips', 'os_type', 'agent_enabled',
            'template', 'protection', 'tags', 'description'
        ]
        
        # Si es archivo nuevo o no tiene headers correctos, escribir headers
        if not existing_file or ws.max_row < 1 or ws.cell(row=1, column=1).value != 'fecha_actualizacion':
            print("🔧 Configurando headers...")
            # Escribir headers con formato
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            print("✓ Headers existentes mantenidos")
        
        # LIMPIAR DATOS: Eliminar todas las filas desde la 2 hacia abajo
        if ws.max_row > 1:
            print(f"🧹 Limpiando {ws.max_row - 1} filas de datos antiguos...")
            ws.delete_rows(2, ws.max_row - 1)
        
        # Escribir nuevos datos empezando desde la fila 2
        print(f"📝 Escribiendo {len(inventory)} registros nuevos...")
        for row, vm in enumerate(inventory, 2):
            for col, header in enumerate(headers, 1):
                value = vm.get(header, 'N/A')
                
                # Formatear valores
                if header in ['memory_gb', 'total_disk_gb'] and isinstance(value, (int, float)):
                    value = round(float(value), 2)
                elif header in ['cpu_cores', 'cpu_sockets', 'disk_count', 'network_interfaces', 'ip_count']:
                    value = int(value) if str(value).isdigit() else 0
                
                cell = ws.cell(row=row, column=col, value=value)
                
                # Formato especial para fecha
                if header == 'fecha_actualizacion':
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    cell.font = Font(bold=True)
                
                # Colores por estado
                elif header == 'status':
                    if value == 'running':
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif value == 'stopped':
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Solo ajustar columnas si es archivo nuevo
        if not existing_file:
            print("📏 Ajustando ancho de columnas...")
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        print("✅ Excel actualizado en memoria")
        return excel_buffer
    
    def create_excel_in_memory(self, inventory):
        """Crear archivo Excel en memoria (función legacy - usar update_existing_excel_in_memory)"""
        if not SHAREPOINT_AVAILABLE:
            return BytesIO()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VM Inventory"
        
        if not inventory:
            return BytesIO()
        
        # Headers con fecha de actualización al inicio
        headers = [
            'fecha_actualizacion', 'cluster_name', 'cluster_type', 'node', 'vmid', 'name', 'status',
            'cpu_cores', 'cpu_sockets', 'memory_gb', 'total_disk_gb', 'disk_count',
            'network_interfaces', 'ip_count', 'ips', 'os_type', 'agent_enabled',
            'template', 'protection', 'tags', 'description'
        ]
        
        # Escribir headers con formato
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir datos
        for row, vm in enumerate(inventory, 2):
            for col, header in enumerate(headers, 1):
                value = vm.get(header, 'N/A')
                
                # Formatear valores
                if header in ['memory_gb', 'total_disk_gb'] and isinstance(value, (int, float)):
                    value = round(float(value), 2)
                elif header in ['cpu_cores', 'cpu_sockets', 'disk_count', 'network_interfaces', 'ip_count']:
                    value = int(value) if str(value).isdigit() else 0
                
                cell = ws.cell(row=row, column=col, value=value)
                
                # Formato especial para fecha
                if header == 'fecha_actualizacion':
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    cell.font = Font(bold=True)
                
                # Colores por estado
                elif header == 'status':
                    if value == 'running':
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif value == 'stopped':
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Ajustar columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def upload_to_sharepoint(self, inventory, folder_path=None, filename=None):
        """Subir archivo a carpeta específica actualizando Excel existente"""
        if not self.is_configured():
            print("❌ SharePoint no configurado o dependencias faltantes")
            return False
        
        if not self.access_token:
            if not self.authenticate():
                return False
        
        # Usar valores por defecto del .env si no se especifican
        folder_path = folder_path or self.default_folder
        filename = filename or self.default_filename
        
        # Asegurar que siempre esté dentro de carpeta Inventory
        if not folder_path.startswith('Inventory/'):
            folder_path = f"Inventory/{folder_path}"
        
        try:
            site_id = self.get_site_id()
            if not site_id:
                return False
            
            # Crear carpeta Inventory si no existe
            self.create_folder_if_not_exists(site_id, "Inventory")
            
            # Crear subcarpeta si es necesaria
            if "/" in folder_path and folder_path != "Inventory":
                self.create_folder_if_not_exists(site_id, folder_path)
            
            # Construir ruta completa considerando la biblioteca
            if self.library_name and self.library_name != 'Documents':
                full_path = f"{self.library_name}/{folder_path}/{filename}"
            else:
                full_path = f"{folder_path}/{filename}"
            
            # NUEVA LÓGICA: Actualizar Excel existente en lugar de crear nuevo
            excel_data = self.update_existing_excel_in_memory(inventory, site_id, full_path)
            
            upload_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{full_path}:/content"
            
            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            print(f"📤 Actualizando archivo: {full_path}")
            response = requests.put(upload_url, headers=headers, data=excel_data.getvalue())
            
            if response.status_code in [200, 201]:
                result = response.json()
                web_url = result.get('webUrl', 'URL no disponible')
                
                print(f"✅ Archivo actualizado exitosamente en SharePoint!")
                print(f"🔄 Datos actualizados preservando estructura existente")
                if self.library_name != 'Documents':
                    print(f"📚 Biblioteca: {self.library_name}")
                print(f"📁 Ruta completa: {folder_path}")
                print(f"📎 URL del archivo: {web_url}")
                print(f"📄 Nombre: {filename}")
                print(f"🕒 Última actualización: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                print(f"📊 Total de VMs: {len(inventory)}")
                
                running_vms = len([vm for vm in inventory if vm.get('status') == 'running'])
                stopped_vms = len([vm for vm in inventory if vm.get('status') == 'stopped'])
                print(f"🟢 VMs en ejecución: {running_vms}")
                print(f"🔴 VMs detenidas: {stopped_vms}")
                print(f"🔗 Las tablas dinámicas existentes seguirán funcionando")
                
                return True
            
            # Manejo específico de errores
            elif response.status_code == 423:
                error_data = response.json()
                print(f"🔒 Archivo bloqueado en SharePoint")
                print(f"💡 Solución: Cierra el archivo '{filename}' en SharePoint y vuelve a intentar")
                print(f"📝 El archivo está siendo editado por alguien o abierto en Excel Online")
                print(f"⏰ Espera unos minutos y vuelve a ejecutar el comando")
                return False
            
            elif response.status_code == 404:
                print(f"❌ Ruta no encontrada: {full_path}")
                print(f"💡 Verifica la configuración del sitio SharePoint:")
                print(f"   - SHAREPOINT_URL: {self.sharepoint_url}")
                print(f"   - SHAREPOINT_SITE_NAME: {self.site_name}")
                print(f"   - SHAREPOINT_LIBRARY_NAME: {self.library_name}")
                return False
            
            elif response.status_code == 401:
                print(f"❌ Error de autenticación con SharePoint")
                print(f"💡 Verifica las credenciales en tu archivo .env:")
                print(f"   - SHAREPOINT_TENANT_ID")
                print(f"   - SHAREPOINT_CLIENT_ID") 
                print(f"   - SHAREPOINT_CLIENT_SECRET")
                return False
            
            elif response.status_code == 403:
                print(f"❌ Sin permisos para escribir en SharePoint")
                print(f"💡 Verifica que la aplicación Azure AD tenga permisos de escritura")
                print(f"📂 Ruta intentada: {full_path}")
                return False
            
            else:
                print(f"❌ Error actualizando archivo: {response.status_code}")
                try:
                    error_data = response.json()
                    error_message = error_data.get('error', {}).get('message', 'Error desconocido')
                    print(f"📝 Mensaje: {error_message}")
                except:
                    print(f"📝 Respuesta: {response.text}")
                return False
                
        except Exception as e:
            print(f"❌ Error uploading to SharePoint: {e}")
            return False
