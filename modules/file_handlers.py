"""
Módulo para manejar la exportación de archivos (CSV, JSON, Excel)
"""

import os
import json
import csv
from datetime import datetime

# Imports opcionales para Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl no instalado. Funcionalidad Excel deshabilitada.")


class FileExporter:
    """Manejador de exportación de archivos"""
    
    def __init__(self, output_dir="outputs"):
        self.output_dir = output_dir
        self._ensure_output_directories()
    
    def _ensure_output_directories(self):
        """Crear directorios de salida si no existen"""
        directories = [
            self.output_dir,
            os.path.join(self.output_dir, "excel"),
            os.path.join(self.output_dir, "csv"),
            os.path.join(self.output_dir, "json")
        ]
        
        for directory in directories:
            if not os.path.exists(directory):
                os.makedirs(directory)
                print(f"📁 Carpeta '{directory}' creada")
    
    def _generate_filename(self, format_type):
        """Generar nombre de archivo con timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = f"proxmox_multi_cluster_inventory_{timestamp}"
        
        format_mapping = {
            'csv': ('csv', 'csv'),
            'json': ('json', 'json'),
            'excel': ('excel', 'xlsx')
        }
        
        if format_type in format_mapping:
            folder, extension = format_mapping[format_type]
            return os.path.join(self.output_dir, folder, f"{base_filename}.{extension}")
        
        return None
    
    def save_csv(self, inventory, custom_filepath=None):
        """Guardar inventario en formato CSV"""
        try:
            filepath = custom_filepath or self._generate_filename('csv')
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                if inventory:
                    writer = csv.DictWriter(f, fieldnames=inventory[0].keys())
                    writer.writeheader()
                    writer.writerows(inventory)
            
            print(f"✓ Inventario CSV guardado en: {filepath}")
            return filepath
        except Exception as e:
            print(f"❌ Error guardando CSV: {e}")
            return None
    
    def save_json(self, inventory, custom_filepath=None):
        """Guardar inventario en formato JSON"""
        try:
            filepath = custom_filepath or self._generate_filename('json')
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(inventory, f, indent=2, ensure_ascii=False, default=str)
            
            print(f"✓ Inventario JSON guardado en: {filepath}")
            return filepath
        except Exception as e:
            print(f"❌ Error guardando JSON: {e}")
            return None
    
    def save_excel(self, inventory, custom_filepath=None):
        """Guardar inventario en formato Excel con formato profesional"""
        if not EXCEL_AVAILABLE:
            print("❌ Excel no disponible. Instala openpyxl: pip install openpyxl")
            return None
        
        try:
            filepath = custom_filepath or self._generate_filename('excel')
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "VM Inventory"
            
            if not inventory:
                print("⚠️ No hay datos para guardar en Excel")
                return None
            
            # Headers con fecha de actualización al inicio
            headers = [
                'fecha_actualizacion', 'cluster_name', 'cluster_type', 'node', 'vmid', 'name', 'status',
                'cpu_cores', 'cpu_sockets', 'memory_gb', 'total_disk_gb', 'disk_count',
                'network_interfaces', 'ip_count', 'ips', 'os_type', 'agent_enabled',
                'template', 'protection', 'tags', 'description'
            ]
            
            # Escribir headers con formato
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Escribir datos
            for row, vm in enumerate(inventory, 2):
                for col, header in enumerate(headers, 1):
                    value = vm.get(header, 'N/A')
                    
                    # Formatear valores
                    if header in ['memory_gb', 'total_disk_gb'] and isinstance(value, (int, float)):
                        value = round(float(value), 2)
                    elif header in ['cpu_cores', 'cpu_sockets', 'disk_count', 'network_interfaces', 'ip_count']:
                        value = int(value) if str(value).isdigit() else 0
                    
                    cell = ws.cell(row=row, column=col, value=value)
                    
                    # Formato especial para fecha
                    if header == 'fecha_actualizacion':
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        cell.font = Font(bold=True)
                    
                    # Colores por estado
                    elif header == 'status':
                        if value == 'running':
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif value == 'stopped':
                            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Ajustar columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Agregar información de última actualización
            last_update_info = f"Última actualización: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A' + str(len(inventory) + 3)] = last_update_info
            ws['A' + str(len(inventory) + 3)].font = Font(bold=True, size=12)
            ws['A' + str(len(inventory) + 3)].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            
            wb.save(filepath)
            print(f"✓ Inventario Excel guardado en: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"❌ Error guardando Excel: {e}")
            return None
    
    def save_all_formats(self, inventory):
        """Guardar inventario en todos los formatos disponibles"""
        if not inventory:
            print("❌ No hay datos para guardar")
            return {}
        
        results = {}
        
        print(f"\n📁 Guardando inventario en múltiples formatos...")
        
        # CSV
        csv_path = self.save_csv(inventory)
        if csv_path:
            results['csv'] = csv_path
        
        # JSON
        json_path = self.save_json(inventory)
        if json_path:
            results['json'] = json_path
        
        # Excel
        if EXCEL_AVAILABLE:
            excel_path = self.save_excel(inventory)
            if excel_path:
                results['excel'] = excel_path
        
        return results
    
    def save_by_format(self, inventory, output_format):
        """Guardar inventario según el formato especificado"""
        if not inventory:
            print("❌ No hay datos para guardar")
            return None
        
        if output_format == 'csv':
            return self.save_csv(inventory)
        elif output_format == 'json':
            return self.save_json(inventory)
        elif output_format == 'excel':
            return self.save_excel(inventory)
        elif output_format == 'all':
            return self.save_all_formats(inventory)
        else:
            print(f"❌ Formato no soportado: {output_format}")
            return None
